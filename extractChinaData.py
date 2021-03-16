import csv
import collections
from openpyxl import Workbook

def extract_data(file):
    province_data = collections.defaultdict(list)
    with open(file, 'r') as f:
        lines = csv.reader(f)
        for line in lines:
            province_name = line[0]
            country_name = line[1]
            data_length = len(line)
            if country_name == 'China':
                for i in range(4, data_length):
                    province_data[province_name].append(int(line[i]))
    return province_data

confirmed_data = extract_data('global_cv_data/time_series_covid19_confirmed_global.csv')
death_data = extract_data('global_cv_data/time_series_covid19_deaths_global.csv')
recovered_data = extract_data('global_cv_data/time_series_covid19_recovered_global.csv')

provinceList = list(confirmed_data.keys())

ChinaPopulation = 1404676330
row_lenght = len(confirmed_data['Anhui'])

ChinaConfirmed = []
ChinaDeath = []
ChinaRecovered = []
for i in range(row_lenght):
    cData = dData = rData = 0
    for p in provinceList:
        cData += confirmed_data[p][i]
        dData += death_data[p][i]
        rData += recovered_data[p][i]
    ChinaConfirmed.append(cData)
    ChinaDeath.append(dData)
    ChinaRecovered.append(rData)
    
ChinaInfected = []
ChinaRemoved = []
for i in range(len(ChinaConfirmed)):
    ChinaInfected.append(ChinaConfirmed[i]-ChinaDeath[i]-ChinaRecovered[i])
    ChinaRemoved.append(ChinaDeath[i]+ChinaRecovered[i])

wb = Workbook()
ws = wb.active
cd = ChinaConfirmed
dd = ChinaDeath
rd = ChinaRecovered
id = ChinaInfected
rmd = ChinaRemoved
po = ChinaPopulation
dl = len(cd)
ws.cell(row = 1, column = 1, value = 'confirmed')
ws.cell(row = 1, column = 2, value = 'death')
ws.cell(row = 1, column = 3, value = 'recovered')
ws.cell(row = 1, column = 4, value = 'infected')
ws.cell(row = 1, column = 5, value = 'population')
ws.cell(row = 1, column = 6, value = 'susceptible')
ws.cell(row = 1, column = 7, value = 'rate')
ws.cell(row = 1, column = 8, value = 'removed')
for i in range(dl):
    ws.cell(row = i+2, column = 1, value = cd[i])
    ws.cell(row = i+2, column = 2, value = dd[i])
    ws.cell(row = i+2, column = 3, value = rd[i])
    ws.cell(row = i+2, column = 4, value = id[i])
    ws.cell(row = i+2, column = 5, value = po)
    susceptible = po-cd[i]
    ws.cell(row = i+2, column = 6, value = susceptible)
    ws.cell(row = i+2, column = 7, value = susceptible/po)
    ws.cell(row = i+2, column = 8, value = rmd[i])
wb.save('cv_data_with_el/%s.xlsx'%'China')