# -*- coding='utf-8' -*-

# import rpy2
import csv
import collections
from openpyxl import Workbook
# import rpy2.robjects as robjects
# from rpy2.robjects.packages import importr

def extract_data(file):
    country_data = collections.defaultdict(list)
    with open(file, 'r') as f:
        lines = csv.reader(f)
        for line in lines:
            province_name = line[0]
            country_name = line[1]
            data_length = len(line)
            if province_name == '':
                for i in range(4, data_length):
                    country_data[country_name].append(int(line[i]))
    return country_data

confirmed_data = extract_data('global_cv_data/time_series_covid19_confirmed_global.csv')
death_data = extract_data('global_cv_data/time_series_covid19_deaths_global.csv')
recovered_data = extract_data('global_cv_data/time_series_covid19_recovered_global.csv')

country_list = list(confirmed_data.keys())
country_list.sort()

infected_data = collections.defaultdict(list)
removed_data = collections.defaultdict(list)
for Cname in country_list:
    if confirmed_data[Cname] and death_data[Cname] and recovered_data[Cname]:
        cd = confirmed_data[Cname]
        dd = death_data[Cname]
        rd = recovered_data[Cname]
        dl = len(cd)
        for i in range(dl):
            infected_data[Cname].append(cd[i]-dd[i]-rd[i])
            removed_data[Cname].append(dd[i]+rd[i])

country_population = {}
with open('global_cv_data/UID_ISO_FIPS_LookUp_Table.csv', 'r') as f:
    next(f)
    lines = csv.reader(f)
    for line in lines:
        country_name = line[7]
        if country_name in country_list and line[6] == '':
            try:
                country_population[country_name] = int(line[11])
            except:
                pass

def export_cv_data(Cname):
    wb = Workbook()
    ws = wb.active
    cd = confirmed_data[Cname]
    dd = death_data[Cname]
    rd = recovered_data[Cname]
    id = infected_data[Cname]
    rmd = removed_data[Cname]
    po = country_population[Cname]
    dl = len(cd)
    ws.cell(row = 1, column = 1, value = 'confirmed')
    ws.cell(row = 1, column = 2, value = 'death')
    ws.cell(row = 1, column = 3, value = 'recovered')
    ws.cell(row = 1, column = 4, value = 'infected')
    ws.cell(row = 1, column = 5, value = 'population')
    ws.cell(row = 1, column = 6, value = 'susceptible')
    ws.cell(row = 1, column = 7, value = 'rate')
    ws.cell(row = 1, column = 8, value = 'removed')
    for i in range(dl):
        ws.cell(row = i+2, column = 1, value = cd[i])
        ws.cell(row = i+2, column = 2, value = dd[i])
        ws.cell(row = i+2, column = 3, value = rd[i])
        ws.cell(row = i+2, column = 4, value = id[i])
        ws.cell(row = i+2, column = 5, value = po)
        susceptible = po-cd[i]
        ws.cell(row = i+2, column = 6, value = susceptible)
        ws.cell(row = i+2, column = 7, value = susceptible/po)
        ws.cell(row = i+2, column = 8, value = rmd[i])
    wb.save('cv_data_with_el/%s.xlsx'%Cname)

for Cname in country_list:
    try:
        export_cv_data(Cname)
    except:
        pass